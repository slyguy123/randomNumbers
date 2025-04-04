import os
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def map_odd_even_and_save(csv_file, output_file):
    """
    Reads a CSV file containing number sets, maps odd numbers to '-' and even numbers to '*',
    and writes the results to an Excel file with colors.
    """
    mapped_data = []

    with open(csv_file, 'r') as file:
        reader = csv.reader(file)
        header = next(reader, None)  # Read header if exists
        for row in reader:
            mapped_row = ['*' if int(num.strip()) % 2 == 0 else '-' for num in row[1].split(',')]
            mapped_data.append(mapped_row)

    # Convert to DataFrame
    df = pd.DataFrame(mapped_data)
    
    # Save as Excel file
    excel_file = output_file.replace(".csv", ".xlsx")
    df.to_excel(excel_file, index=False, header=False)

    # Open with openpyxl to apply colors
    wb = Workbook()
    ws = wb.active

    # Define color fills
    even_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Blue for even
    odd_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Green for odd

    # Write data with colors
    for i, row in enumerate(mapped_data, start=1):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.fill = even_fill if value == '*' else odd_fill

    wb.save(excel_file)
    print(f"Mapped numbers saved to {excel_file}")

if __name__ == "__main__":
    curDir = os.getcwd()
    fName = "euromillions_results.csv"
    csv_file = os.path.join(curDir, fName)
    output_file = os.path.join(curDir, "mapped_numbers.csv")

    map_odd_even_and_save(csv_file, output_file)
